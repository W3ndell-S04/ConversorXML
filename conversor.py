import tkinter as tk
from tkinter import filedialog, messagebox
import xml.etree.ElementTree as ET
from fpdf import FPDF
from openpyxl import Workbook

class XMLConverterApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Conversor XML")
        self.root.geometry("400x250") 

        self.label_xml = tk.Label(root, text="Arquivo XML selecionado:")
        self.label_xml.pack()

        self.selected_xml_label = tk.Label(root, text="")
        self.selected_xml_label.pack()

        self.btn_xml = tk.Button(root, text="Selecionar XML", command=self.select_xml_file)
        self.btn_xml.pack()

        self.label_output = tk.Label(root, text="Pasta de saída:")
        self.label_output.pack()

        self.selected_output_label = tk.Label(root, text="")
        self.selected_output_label.pack()

        self.btn_output = tk.Button(root, text="Selecionar Pasta", command=self.select_output_location)
        self.btn_output.pack()

        self.save_option = tk.StringVar()
        self.save_option.set("pdf")  # Valor padrão

        self.pdf_radio = tk.Radiobutton(root, text="Salvar como PDF", variable=self.save_option, value="pdf")
        self.pdf_radio.pack()

        self.excel_radio = tk.Radiobutton(root, text="Salvar como Excel", variable=self.save_option, value="excel")
        self.excel_radio.pack()

        self.btn_convert = tk.Button(root, text="Converter", command=self.convert)
        self.btn_convert.pack()

    def select_xml_file(self):
        self.xml_path = filedialog.askopenfilename(title="Selecione o arquivo XML")
        self.selected_xml_label.config(text=self.xml_path)
    
    def select_output_location(self):
        self.output_folder = filedialog.askdirectory(title="Selecione a pasta de saída")
        self.selected_output_label.config(text=self.output_folder)

    def convert(self):
        if not hasattr(self, 'xml_path') or not hasattr(self, 'output_folder'):
            messagebox.showerror("Erro", "Por favor, selecione tanto o arquivo XML quanto a pasta de saída.")
            return

        informacoes_nota = self.extract_information_from_xml(self.xml_path)
        if self.save_option.get() == "pdf":
            pdf_path = self.output_folder + "/nota_fiscal.pdf"
            self.create_pdf(informacoes_nota, pdf_path)
            messagebox.showinfo("Conversão Concluída", "A nota fiscal foi convertida para PDF com sucesso!")
        elif self.save_option.get() == "excel":
            excel_path = self.output_folder + "/nota_fiscal.xlsx"
            self.create_excel(informacoes_nota, excel_path)
            messagebox.showinfo("Conversão Concluída", "A nota fiscal foi convertida para Excel com sucesso!")

    def extract_information_from_xml(self, xml_path):
        informacoes = {}
        tree = ET.parse(xml_path)
        root = tree.getroot()
        
        for child in root:
            if child.tag == 'produtos':
                informacoes[child.tag] = []
                for produto in child:
                    produto_info = {}
                    for info in produto:
                        produto_info[info.tag] = info.text
                    informacoes[child.tag].append(produto_info)
            else:
                informacoes[child.tag] = child.text

        return informacoes

    def create_pdf(self, informacoes, pdf_path):
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", size=12)

        for chave, valor in informacoes.items():
            if isinstance(valor, list):
                pdf.cell(200, 10, txt=f"{chave}:", ln=True, align='L')
                for produto in valor:
                    for produto_chave, produto_valor in produto.items():
                        pdf.cell(200, 10, txt=f"{produto_chave}: {produto_valor}", ln=True, align='L')
            else:
                pdf.cell(200, 10, txt=f"{chave}: {valor}", ln=True, align='L')

        pdf.output(pdf_path)

    def create_excel(self, informacoes, excel_path):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Nota Fiscal"

        for chave, valor in informacoes.items():
            if isinstance(valor, list):
                worksheet.append([f"{chave}:"])
                for produto in valor:
                    for produto_chave, produto_valor in produto.items():
                        worksheet.append([f"{produto_chave}", f"{produto_valor}"])
            else:
                worksheet.append([f"{chave}", f"{valor}"])

        workbook.save(excel_path)

def main():
    root = tk.Tk()
    app = XMLConverterApp(root)
    root.mainloop()

if __name__ == "__main__":
    main()
